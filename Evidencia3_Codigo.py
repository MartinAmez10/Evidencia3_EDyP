import csv
import os
import openpyxl
import sqlite3
from sqlite3 import Error 
import datetime

# Testing
separador = "*"*100

datos_grabar = dict()
datos_leer = dict()

datos = {}
datos_autor = {}
datos_genero = {}

año_actual = datetime.datetime.now().year

# Aquí reviso si existe un archivo llamado biblioteca.db que es donde aparece la base de datos
if not os.path.exists("biblioteca.db"):
  # Si NO existe, se avisa al usuario y muestra un mensaje de que se está creando
  print('No se ha encontrado una versión de datos previa. Se procede a crear un almacén de datos a continuación')
  print('\tGenerando almacén de datos . . .')
  
  # Aquí creamos el almacén, nos conectamos a la base de datos y creamos una tabla ya que no existen     
  with sqlite3.connect("biblioteca.db") as conn:
    bi_cursor = conn.cursor()
    bi_cursor.execute("CREATE TABLE IF NOT EXISTS GENERO (Id_gen INTEGER PRIMARY KEY, nomGen TEXT NOT NULL);")
    bi_cursor.execute("CREATE TABLE IF NOT EXISTS AUTOR (Id_autor INTEGER PRIMARY KEY, apAutor TEXT NOT NULL, nomAutor TEXT NOT NULL);")
    bi_cursor.execute("CREATE TABLE IF NOT EXISTS BIBLIOTECA (Id_libro INT PRIMARY KEY NOT NULL, titulo VARCHAR(32) NOT NULL, AUTOR INTEGER NOT NULL, GENERO INTEGER NOT NULL, año_publicado TEXT NOT NULL, ISBN VARCHAR2(13) NOT NULL, fecha_adquirido TEXT NOT NULL, FOREIGN KEY(AUTOR) REFERENCES AUTOR(Id_autor), FOREIGN KEY(GENERO) REFERENCES GENERO(Id_gen));")
    
    print('AVISO:\t¡Almacén generado con éxito!\n')
else:
  # En caso de que SI exista un almacén de datos llamado así, simplemente hacemos una conexión a la base de datos
  with sqlite3.connect("biblioteca.db") as conn:
    bi_cursor = conn.cursor()        

while True:
  print("Hola! selecciona una opcion que quieras realizar (escribe el numero):")
  print("[1]- Registrar nuevo ejemplar")
  print("[2]- Consultas y Reportes")
  print("[3]- Agregar autor")
  print("[4]- Agregar Genero")
  print("[5]- Salir")
  op_main = int(input())
  
  if op_main == 1:

    with sqlite3.connect("biblioteca.db") as conn:
      bi_cursor = conn.cursor()
      bi_cursor.execute("SELECT * FROM AUTOR, GENERO")
      registro_datos = bi_cursor.fetchall()

    if registro_datos:
        
      # Registro de nuevo ejemplar
      while True:
        autor_nombre = ""
        autor_apellidos = ""
        genero_nombre = ""
        autor_id_evaluado = 0
        genero_id_evaluado = 0

        identificador = max(datos, default=0) + 1
        titulo = input("Dame el nombre del libro: \n").upper()

        # Leer autores de la base de datos
        bi_cursor.execute("SELECT Id_autor, nomAutor, apAutor FROM AUTOR")
        autores = bi_cursor.fetchall()
        print("Autores disponibles:")
        for autor in autores:
          print(f"{autor[0]}. {autor[1]} {autor[2]}")

        while True:
          try:
            autor_id_evaluado = int(input(f"Selecciona el ID del autor para {titulo}: "))
            if autor_id_evaluado in [autor[0] for autor in autores]:
              break
            else:
              print("El ID ingresado no es válido. Ingresa un ID válido.")
          except ValueError:
            print("El ID ingresado no es válido. Ingresa un ID válido.")

        # Leer géneros de la base de datos
        bi_cursor.execute("SELECT Id_gen, nomGen FROM GENERO")
        generos = bi_cursor.fetchall()
        print("Géneros disponibles:")
        for genero in generos:
          print(f"{genero[0]}. {genero[1]}")

        while True:
          try:
            genero_id_evaluado = int(input(f"Selecciona el ID del género para {titulo}: "))
            if genero_id_evaluado in [genero[0] for genero in generos]:
              break
            else:
              print("ID de género inválido. Ingresa un ID válido.")
          except ValueError:
            print("ID de género inválido. Ingresa un ID válido.")

        # Obtener nombre y apellidos del autor seleccionado
        for autor in autores:
          if autor[0] == autor_id_evaluado:
            autor_nombre = autor[1]
            autor_apellidos = autor[2]
            break

        # Obtener nombre del género seleccionado
        for genero in generos:
          if genero[0] == genero_id_evaluado:
            genero_nombre = genero[1]
            break

        while True:
          año_publicacion = int(input(f"En qué año se publicó {titulo}: \n"))
          if año_publicacion <= año_actual:
            break
          else:
            print('VALUE ERROR: Año fuera de los parámetros permitidos. Ingresar un año válido a la fecha actual.')

        ISBN = input(f"Cual es el ISBN de {titulo}: \n").upper()

        while True:
          try:
            fecha_adquisicion = input("Cuándo se adquirió el libro (En formato DD/MM/YYYY): \n")
            fecha_adquisicion = datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y").date()
            break
          except ValueError:
            print("La fecha capturada no es válida. Vuelve a ingresar la fecha en el formato indicado.")

        datos[identificador] = [titulo, f"{autor_nombre} {autor_apellidos}", genero_nombre, año_publicacion, ISBN, fecha_adquisicion]
        print("Datos cargados!")

        with sqlite3.connect("biblioteca.db") as conn:
          bi_cursor = conn.cursor()
          # Ingresamos estos datos en la base de datos generada
          valores_ejemplar = (identificador, titulo, autor_id_evaluado, genero_id_evaluado, año_publicacion, ISBN, fecha_adquisicion)
          bi_cursor.execute("INSERT INTO BIBLIOTECA VALUES(?,?,?,?,?,?,?)", valores_ejemplar)

        op_registro = input("¿Deseas agregar más? (Presiona Enter para no agregar más)\n")
        if op_registro.strip() == "":
          break

      print("")
    else:
      print('\tERROR: Favor de registrar un autor o género previamente')
      print('\nRegresando al menú principal . . .')
      esperar = input('Presione enter para continuar')

  elif op_main == 2:
    with sqlite3.connect("biblioteca.db") as conn:
      bi_cursor = conn.cursor()
      bi_cursor.execute("SELECT titulo, año_publicado FROM BIBLIOTECA")
      registro_datos = bi_cursor.fetchall()

      #bi_cursor.execute("SELECT titulo, año_publicado FROM BIBLIOTECA")
      #regDat = bi_cursor.fetchall()
      #if regDat:
        #for titulo, año_publicado in regDat:
          #print(f'{titulo}\t{año_publicado}')
      #else:
        #print('No se guardan los datos')

    if registro_datos:
        # Consultas y Reportes
      while True:
        print("Selecciona una opcion que quieras realizar (escribe el numero):")
        print("[1]- Consulta de titulo")
        print("[2]- Reportes")
        print("[3]- Regresar al menu principal")
        op_consulta = int(input())
        
        # Separamos por la opción seleccionada
        if op_consulta == 1:
          # Consulta de título o ISBN
          while True:
            print("De que forma deseas buscar el libro?(escribe el numero):")
            print("[1]- Por titulo")
            print("[2]- Por ISBN")
            print("[3]- Regresar al menu anterior")
            op_busqueda = int(input())
            if op_busqueda == 1:
              # Muestra el catálago de Libros (POR TÍTULO)
              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                bi_cursor.execute("SELECT titulo FROM BIBLIOTECA")
                registro_titulo = bi_cursor.fetchall()
              if registro_titulo:
                print('\nTítulos')
                print(separador)
                for titulo in registro_titulo:
                  print(f'{titulo}')
              # Añadí esto para filtrar por título y mostrar información (Maybe lo módifico)
              titulo_buscar = input('Seleccione el título a mostrar: ').upper()
              valor_titulo = {"titulo":titulo_buscar}
              bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR WHERE BIBLIOTECA.titulo = :titulo", valor_titulo)
              registro_titulo_imprimir = bi_cursor.fetchall()

              if registro_titulo_imprimir:
                print('Titulo\t\tAutor\t\t\tGenero\tAño Publicado\tISBN\tFecha Adquirido')
                print(separador)
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_titulo_imprimir:
                  print(f'{titulo}\t{nomAutor} {apAutor}\t{nomGen}\t{año_publicado}\t{ISBN}\t{fecha_adquirido}')
              else:
                print('NO se encontró ningún libro con ese título')

              print(' ')
            elif op_busqueda == 2:
              # Muestra el libro (POR ISBN)
              isbn_buscar = input('Ingrese el ISBN: ')
              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                valores_isbn = {"ISBN": isbn_buscar}
                bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR WHERE BIBLIOTECA.ISBN = :ISBN", valores_isbn)
                registro_isbn = bi_cursor.fetchall()

              if registro_isbn:
                print('Titulo\t\tAutor\t\t\tGenero\tAño Publicado\tISBN\tFecha Adquirido')
                print(separador)
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_isbn:
                  print(f'{titulo}\t{nomAutor} {apAutor}\t{nomGen}\t{año_publicado}\t{ISBN}\t{fecha_adquirido}')
              else:
                print('ERROR: No se encontraron datos, asegurese de haber ingresado correctamente el ISBN')

                        
            elif op_busqueda == 3:
              break
        elif op_consulta == 2:
          # Reportes tabulados
          while True:
            print("Escoge una forma de filtrar los datos:")
            print("[1]- Catálogo completo")
            print("[2]- Por autor")
            print("[3]- Por genero")
            print("[4]- Por año de publicacion")
            print("[5]- Regresar al menu anterior")
            op_reporte = int(input())
            if op_reporte == 1:
              # Catálogo completo
              #datos_grabar = dict()
              print("DATOS GUARDADOS:")
              print('TITULO', ' '*29, 'AUTOR', ' '*18, 'GÉNERO', ' '*8, 'AÑO', ' '*5, 'ISBN', ' '*8, 'ADQUIRIDO   ')
              print(separador)

              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                #bi_cursor.execute("SELECT BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA LEFT JOIN GENERO ON BIBLIOTECA.Id_gen = GENERO.Id_gen LEFT JOIN AUTOR on BIBLIOTECA.Id_autor = AUTOR.Id_autor")
                bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR")
                registro_catCompleto = bi_cursor.fetchall()
              if registro_catCompleto:
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_catCompleto:
                  print(f'{titulo}\t\t\t{nomAutor} {apAutor}\t\t{nomGen}\t{año_publicado}\t{ISBN}\t\t{fecha_adquirido}')
             

              # Exportación a formatos CSV o MsExcel
              print("Desea exportar los datos a algun formato de los siguientes?")
              print("[1]- CSV")
              print("[2]- MsExcel")
              print("[3]- Ninguno")
              op_exportar = int(input())

              # exportacion a CSV o MsExcel
              if op_exportar == 1:
                
                #for j in datos:
                  #datos_grabar = datos
                archivo = open('logs_catalogo_completo.csv', 'w', newline='')
                grabador = csv.writer(archivo)
                grabador.writerow(('identificador', 'titulo', 'autor', 'apellidos', 'genero', 'año_publicacion', 'ISBN', 'fecha_adquisicion'))
                print("Se exporto correctamente!")
                #grabador.writerows([(identificador, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for identificador, datos in datos_grabar.items()])
                grabador.writerows(registro_catCompleto)
                archivo.close()
                #'identificador, titulo, autor, genero, año_publicacion, ISBN, fecha_adquisicion'
                break
            
              # exportacion a MsExcel
              if op_exportar == 2:
                libro = openpyxl.Workbook()
                hoja = libro["Sheet"]
                hoja.title = "primera"
                hoja["A1"].value = "Identificador"
                hoja["B1"].value = "Titulo"
                hoja["C1"].value = "Autor"
                hoja["D1"].value = "Apellidos"
                hoja["E1"].value = "Genero"
                hoja["F1"].value = "Año de publicacion"
                hoja["G1"].value = "ISBN"
                hoja["H1"].value = "Fecha adquisicion"

                #fila = 2

                for i in registro_catCompleto:
                  hoja.append(i)
                #for identificador, dato in datos.items():
                  #hoja.cell(row=fila, column=1).value = identificador
                  #hoja.cell(row=fila, column=2).value = datos[identificador][0]
                  #hoja.cell(row=fila, column=3).value = datos[identificador][1]
                  #hoja.cell(row=fila, column=4).value = datos[identificador][2]
                  #hoja.cell(row=fila, column=5).value = datos[identificador][3]
                  #hoja.cell(row=fila, column=6).value = datos[identificador][4]
                  #hoja.cell(row=fila, column=7).value = datos[identificador][5]
                  #fila += 1

                libro.save("reporte_completo.xlsx")
                print("Se exporto de manera correcta")
                break

            elif op_reporte == 2:
              # Filtro por autor
              #datos_grabar = dict()
              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                bi_cursor.execute("SELECT * FROM AUTOR")
                registro_autor = bi_cursor.fetchall()
              if registro_autor:
                print('\nSeleccione la ID de entre los siguientes autores: ')
                print('ID\t\tAutor')
                print(separador)
                for Id_autor, nomAutor, apAutor in registro_autor:
                  print(f'{Id_autor}\t{nomAutor} {apAutor}')
              #print('Seleccione entre los siguientes autores:')
              #for j in datos:
                #print('Seleccione entre las siguientes opciones:')
                #print(datos[j][1])
              filtro_autor = input("Dame el ID del autor: \n").upper()

              print('TITULO', ' '*29, 'AUTOR', ' '*18, 'GÉNERO', ' '*8, 'AÑO', ' '*5, 'ISBN', ' '*8, 'ADQUIRIDO   ')
              valor_autor = {"Id_autor":filtro_autor}
              bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR WHERE AUTOR.Id_autor = :Id_autor", valor_autor)
              registro_autor_imprimir = bi_cursor.fetchall()

              if registro_autor_imprimir:
                print('Titulo\t\tAutor\t\t\tGenero\tAño Publicado\tISBN\tFecha Adquirido')
                print(separador)
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_autor_imprimir:
                  print(f'{titulo}\t{nomAutor} {apAutor}\t{nomGen}\t{año_publicado}\t{ISBN}\t{fecha_adquirido}')
              else:
                print('NO se encontró ningún libro con ese Autor')
              #for i in datos:
                #if filtro_autor == datos[i][1]:
                  #print(f'{datos[i][0]:35} {datos[i][1]:25} {datos[i][2]:15} {datos[i][3]:8} {datos[i][4]:15} {datos[i][5]:12}')
                  #datos_grabar[i] = datos[i][0], datos[i][1], datos[i][2], datos[i][3], datos[i][4], datos[i][5]
                  #print('\n')
              
              # AQUI
                            # exportacion a formato CSV o a MsExcel
              print("Desea exportar los datos a algun formato de los siguientes?")
              print("[1]- CSV")
              print("[2]- MsExcel")
              print("[3]- Ninguno")
              op_exportar = int(input())
              
              # IF para filtrar el formato deseado
              # Exportación a CSV
              if op_exportar == 1:
                #for j in datos:
                  #datos_grabar = datos
                archivo = open('logs_autor_' + filtro_autor.lower() + '.csv', 'w', newline='')
                grabador = csv.writer(archivo)
                grabador.writerow(('identificador', 'titulo', 'autor', 'apellidos', 'genero', 'año_publicacion', 'ISBN', 'fecha_adquisicion'))
                print("Se exporto correctamente!")
                #grabador.writerows([(identificador, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for identificador, datos in datos_grabar.items()])
                grabador.writerows(registro_autor_imprimir)
                archivo.close()
                #'identificador, titulo, autor, genero, año_publicacion, ISBN, fecha_adquisicion'
                break
            
              # Exportación a MsExcel
              if op_exportar == 2:
                libro = openpyxl.Workbook()
                hoja = libro["Sheet"]
                hoja.title = "primera"
                hoja["A1"].value = "Identificador"
                hoja["B1"].value = "Titulo"
                hoja["C1"].value = "Autor"
                hoja["D1"].value = "Apellidos"
                hoja["E1"].value = "Genero"
                hoja["F1"].value = "Año de publicacion"
                hoja["G1"].value = "ISBN"
                hoja["H1"].value = "Fecha adquisicion"

                #fila = 2
                for i in registro_autor_imprimir:
                  hoja.append(i)
                #for identificador, dato in datos.items():
                  #hoja.cell(row=fila, column=1).value = identificador
                  #hoja.cell(row=fila, column=2).value = datos[identificador][0]
                  #hoja.cell(row=fila, column=3).value = datos[identificador][1]
                  #hoja.cell(row=fila, column=4).value = datos[identificador][2]
                  #hoja.cell(row=fila, column=5).value = datos[identificador][3]
                  #hoja.cell(row=fila, column=6).value = datos[identificador][4]
                  #hoja.cell(row=fila, column=7).value = datos[identificador][5]
                  #fila += 1

                libro.save("reporte_autor.xlsx")
                print("Se exporto de manera correcta")
                break
              # AQUI

            elif op_reporte == 3:
              # Filtro por género
              #datos_grabar = dict()
              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                bi_cursor.execute("SELECT * FROM GENERO")
                registro_genero = bi_cursor.fetchall()
              if registro_genero:
                print('\nSeleccione la ID de entre los siguientes generos: ')
                print('ID\t\tGenero')
                print(separador)
                for Id_gen, nomGen in registro_genero:
                  print(f'{Id_gen}\t{nomGen}')
              #print('Seleccione entre los siguientes Géneros:')
              #for j in datos:
                #print('Seleccione entre las siguientes opciones:')
                #print(datos[j][2])
              
              filtro_genero = input("Dame el genero: \n").upper()
              print('TITULO', ' '*29, 'AUTOR', ' '*18, 'GÉNERO', ' '*8, 'AÑO', ' '*5, 'ISBN', ' '*8, 'ADQUIRIDO   ')
              valor_genero = {"Id_gen":filtro_genero}
              bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR WHERE GENERO.Id_gen = :Id_gen", valor_genero)
              registro_genero_imprimir = bi_cursor.fetchall()

              if registro_genero_imprimir:
                print('Titulo\t\tAutor\t\t\tGenero\tAño Publicado\tISBN\tFecha Adquirido')
                print(separador)
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_genero_imprimir:
                  print(f'{titulo}\t{nomAutor} {apAutor}\t{nomGen}\t{año_publicado}\t{ISBN}\t{fecha_adquirido}')
              else:
                print('NO se encontró ningún libro con ese Género')
              #for i in datos:
                #if filtro_genero == datos[i][2]:
                  #print(f'{datos[i][0]:35} {datos[i][1]:25} {datos[i][2]:15} {datos[i][3]:8} {datos[i][4]:15} {datos[i][5]:12}')
                  #datos_grabar[i] = datos[i][0], datos[i][1], datos[i][2], datos[i][3], datos[i][4], datos[i][5]
                  #print('\n')

              # Exportación a formatos CSV o MsExcel
              print("Desea exportar los datos a algun formato de los siguientes?")
              print("[1]- CSV")
              print("[2]- MsExcel")
              print("[3]- Ninguno")
              op_exportar = int(input())

              # IF para filtrar el formato deseado
              # Exportación a CSV
              if op_exportar == 1:
                #for j in datos:
                  #datos_grabar = datos
                archivo = open('logs_genero_' + filtro_genero.lower() + '.csv', 'w', newline='')
                grabador = csv.writer(archivo)
                grabador.writerow(('identificador', 'titulo', 'autor', 'apellidos', 'genero', 'año_publicacion', 'ISBN', 'fecha_adquisicion'))
                print("Se exporto correctamente!")
                #grabador.writerows([(identificador, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for identificador, datos in datos_grabar.items()])
                grabador.writerows(registro_genero_imprimir)
                archivo.close()
                #'identificador, titulo, autor, genero, año_publicacion, ISBN, fecha_adquisicion'
                break
            
              # Exportación a MsExcel
              if op_exportar == 2:
                libro = openpyxl.Workbook()
                hoja = libro["Sheet"]
                hoja.title = "primera"
                hoja["A1"].value = "Identificador"
                hoja["B1"].value = "Titulo"
                hoja["C1"].value = "Autor"
                hoja["D1"].value = "Apellidos"
                hoja["E1"].value = "Genero"
                hoja["F1"].value = "Año de publicacion"
                hoja["G1"].value = "ISBN"
                hoja["H1"].value = "Fecha adquisicion"

                #fila = 2
                for i in registro_genero_imprimir:
                  hoja.append(i)
                #for identificador, dato in datos.items():
                  #hoja.cell(row=fila, column=1).value = identificador
                  #hoja.cell(row=fila, column=2).value = datos[identificador][0]
                  #hoja.cell(row=fila, column=3).value = datos[identificador][1]
                  #hoja.cell(row=fila, column=4).value = datos[identificador][2]
                  #hoja.cell(row=fila, column=5).value = datos[identificador][3]
                  #hoja.cell(row=fila, column=6).value = datos[identificador][4]
                  #hoja.cell(row=fila, column=7).value = datos[identificador][5]
                  #fila += 1

                libro.save("reporte_genero.xlsx")
                print("Se exporto de manera correcta")
                break

            elif op_reporte == 4:
              # Filtrado por año
              #datos_grabar = dict()
              with sqlite3.connect("biblioteca.db") as conn:
                bi_cursor = conn.cursor()
                bi_cursor.execute("SELECT año_publicado FROM BIBLIOTECA")
                registro_año = bi_cursor.fetchall()
              if registro_año:
                print('\nSeleccione entre los siguientes Años de Publicación: ')
                print(separador)
                for año_publicado in registro_año:
                  print(f'{año_publicado}')
              #print('Seleccione entre los siguientes Años de Publicacion:')
              #for j in datos:
                #print(datos[j][3])

              filtro_año = input("Dame el año de publicacion: \n").upper()
              print('TITULO', ' '*29, 'AUTOR', ' '*18, 'GÉNERO', ' '*8, 'AÑO', ' '*5, 'ISBN', ' '*8, 'ADQUIRIDO   ')
              valor_año = {"año_publicado":filtro_año}
              bi_cursor.execute("SELECT BIBLIOTECA.Id_libro, BIBLIOTECA.titulo, AUTOR.nomAutor, AUTOR.apAutor, GENERO.nomGen, BIBLIOTECA.año_publicado, BIBLIOTECA.ISBN, BIBLIOTECA.fecha_adquirido FROM BIBLIOTECA INNER JOIN GENERO ON GENERO.Id_gen = BIBLIOTECA.GENERO INNER JOIN AUTOR ON AUTOR.Id_autor = BIBLIOTECA.AUTOR WHERE GENERO.Id_gen = :Id_gen", valor_genero)
              registro_año_imprimir = bi_cursor.fetchall()

              if registro_año_imprimir:
                print('Titulo\t\tAutor\t\t\tGenero\tAño Publicado\tISBN\tFecha Adquirido')
                print(separador)
                for Id_libro, titulo, nomAutor, apAutor, nomGen, año_publicado, ISBN, fecha_adquirido in registro_año_imprimir:
                  print(f'{titulo}\t{nomAutor} {apAutor}\t{nomGen}\t{año_publicado}\t{ISBN}\t{fecha_adquirido}')
              else:
                print('NO se encontró ningún libro en ese Año de Publicación')
              #for i in datos:
                #if filtro_año == datos[i][3]:
                  #print(f'{datos[i][0]:35} {datos[i][1]:25} {datos[i][2]:15} {datos[i][3]:8} {datos[i][4]:15} {datos[i][5]:12}')
                  #datos_grabar[i] = datos[i][0], datos[i][1], datos[i][2], datos[i][3], datos[i][4], datos[i][5]
                  #print('\n')

              # Exportación a formatos CSV o MsExcel
              print("Desea exportar los datos a algun formato de los siguientes?")
              print("[1]- CSV")
              print("[2]- MsExcel")
              print("[3]- Ninguno")
              op_exportar = int(input())

              # IF para filtrar el formato deseado
              # Exportación a CSV
              if op_exportar == 1:
                #for j in datos:
                  #datos_grabar = datos
                archivo = open('logs_fecha_publicado_' + filtro_año + '.csv', 'w', newline='')
                grabador = csv.writer(archivo)
                grabador.writerow(('identificador', 'titulo', 'autor', 'apellidos', 'genero', 'año_publicacion', 'ISBN', 'fecha_adquisicion'))
                print("Se exporto correctamente!")
                #grabador.writerows([(identificador, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for identificador, datos in datos_grabar.items()])
                grabador.writerows(registro_año_imprimir)
                archivo.close()
                #'identificador, titulo, autor, genero, año_publicacion, ISBN, fecha_adquisicion'
                break
            
              # Exportación a MsExcel
              if op_exportar == 2:
                libro = openpyxl.Workbook()
                hoja = libro["Sheet"]
                hoja.title = "primera"
                hoja["A1"].value = "Identificador"
                hoja["B1"].value = "Titulo"
                hoja["C1"].value = "Autor"
                hoja["D1"].value = "Apellidos"
                hoja["E1"].value = "Genero"
                hoja["F1"].value = "Año de publicacion"
                hoja["G1"].value = "ISBN"
                hoja["H1"].value = "Fecha adquisicion"

                #fila = 2
                for i in registro_año_imprimir:
                  hoja.append(i)
                #for identificador, dato in datos.items():
                  #hoja.cell(row=fila, column=1).value = identificador
                  #hoja.cell(row=fila, column=2).value = datos[identificador][0]
                  #hoja.cell(row=fila, column=3).value = datos[identificador][1]
                  #hoja.cell(row=fila, column=4).value = datos[identificador][2]
                  #hoja.cell(row=fila, column=5).value = datos[identificador][3]
                  #hoja.cell(row=fila, column=6).value = datos[identificador][4]
                  #hoja.cell(row=fila, column=7).value = datos[identificador][5]
                  #fila += 1

                libro.save("reporte_año.xlsx")
                print("Se exporto de manera correcta")
                break
              # AQUI
            elif op_reporte == 5:
              # Regresa al menú anterior
              break
        elif op_consulta == 3:
          # Regresa al menú anterior
          print('Volviendo al menú principal . . .')
          break
    else:
      print('\tERROR: No hay datos que mostrar, favor de registrar un nuevo ejemplar, autor o género previamente')
      print('\nRegresando al menú principal . . .')
      esperar = input('Presione enter para continuar')

  elif op_main == 3:
    # Opción filtrada para: Registrar un Autor
    with sqlite3.connect("biblioteca.db") as conn:
      bi_cursor = conn.cursor()
      bi_cursor.execute("SELECT * FROM AUTOR ")
      bd_autor = bi_cursor.fetchall()

      if bd_autor:
        print("Se han encontrado los siguientes autores registrados:")
        print(separador)
        print("ID\tApellido\tNombre")
        print(separador)
        for Id_autor, apAutor, nomAutor in bd_autor:
          print(f"{Id_autor:^6}\t{apAutor}\t{nomAutor}")

        print(separador)
        print("De los autores anteriormente presentados, desea agregar uno mas?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_autor = int(input())
        if opcion_agregar_autor == 1:
          print("Ingrese los siguientes datos:")
          apellido_autor =  input("Apellido:")
          nombre_autor = input("Nombre: ")

          valores = (apellido_autor, nombre_autor)

          bi_cursor.execute("INSERT INTO AUTOR (apAutor, nomAutor) VALUES(?,?)", valores)
          tabla_autores = bi_cursor.fetchall()
          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

      else:
        print("No se han encontrado autores registrados, desea agregar uno?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_autor = int(input())

        if opcion_agregar_autor == 1:
          print(separador)
          print("ingresa los siguientes datos::")
          apellido_autor =  input("Apellido:")
          nombre_autor = input("Nombre: ")

          valores = (apellido_autor, nombre_autor)

          bi_cursor.execute("INSERT INTO AUTOR (apAutor, nomAutor) VALUES(?,?)", valores)
          tabla_autores = bi_cursor.fetchall()

          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

  elif op_main == 3:
    # Opción filtrada para: Registrar un Autor
    with sqlite3.connect("biblioteca.db") as conn:
      bi_cursor = conn.cursor()
      bi_cursor.execute("SELECT * FROM AUTOR ")
      bd_autor = bi_cursor.fetchall()

      if bd_autor:
        print("Se han encontrado los siguientes autores registrados:")
        print(separador)
        print("ID\tApellido\tNombre")
        print(separador)
        for Id_autor, apAutor, nomAutor in bd_autor:
          print(f"{Id_autor:^6}\t{apAutor}\t{nomAutor}")

        print(separador)
        print("De los autores anteriormente presentados, desea agregar uno mas?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_autor = int(input())
        if opcion_agregar_autor == 1:
          print("Ingrese los siguientes datos:")
          apellido_autor =  input("Apellido:")
          nombre_autor = input("Nombre: ")

          valores = (apellido_autor, nombre_autor)

          bi_cursor.execute("INSERT INTO AUTOR (apAutor, nomAutor) VALUES(?,?)", valores)
          tabla_autores = bi_cursor.fetchall()
          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

      else:
        print("No se han encontrado autores registrados, desea agregar uno?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_autor = int(input())

        if opcion_agregar_autor == 1:
          print(separador)
          print("ingresa los siguientes datos::")
          apellido_autor =  input("Apellido:")
          nombre_autor = input("Nombre: ")

          valores = (apellido_autor, nombre_autor)

          bi_cursor.execute("INSERT INTO AUTOR (apAutor, nomAutor) VALUES(?,?)", valores)
          tabla_autores = bi_cursor.fetchall()

          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

  elif op_main == 4:
    # Opción filtrada para: Registrar un Género
    with sqlite3.connect("biblioteca.db") as conn:
      bi_cursor = conn.cursor()
      bi_cursor.execute("SELECT * FROM GENERO")
      bd_genero = bi_cursor.fetchall()

      if bd_genero:
        print("Se han encontrado los siguientes generos registrados:")
        print(separador)
        print("ID\tGenero")
        print(separador)
        for Id_gen, nomGen in bd_genero:
          print(f"{Id_gen:^6}\t{nomGen}")

        print(separador)
        print("De los generos anteriormente presentados, desea agregar uno mas?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_genero = int(input())

        if opcion_agregar_genero == 1:
          print("Ingrese los siguientes datos:")
          genero_nuevo = input("Genero: ")

          bi_cursor.execute("INSERT INTO GENERO (nomGen) VALUES(?)", (genero_nuevo,))
          tabla_genero = bi_cursor.fetchall()

          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

      else:
        print("No se han encontrado generos registrados, desea agregar uno?:")
        print("[1]- SI")
        print("[2]- NO")
        opcion_agregar_genero = int(input())

        if opcion_agregar_genero == 1:
          print(separador)
          print("ingresa los siguientes datos::")
          genero_nuevo = input("Genero: ")

          bi_cursor.execute("INSERT INTO GENERO (nomGen) VALUES(?)", (genero_nuevo,))
          tabla_genero = bi_cursor.fetchall()

          print("Se cargo correctamente!")
          print(f"clave asignada: {bi_cursor.lastrowid}")
        else:
          print("Regresando al menu principal...")

  elif op_main == 5:
    # Sale del programa
    break

for j in datos:
  datos_grabar = datos
archivo = open('logs_libros.csv', 'w', newline='')

grabador = csv.writer(archivo)
grabador.writerow(('identificador', 'titulo', 'autor', 'genero', 'año_publicacion', 'ISBN', 'fecha_adquisicion'))
#'identificador, titulo, autor, genero, año_publicacion, ISBN, fecha_adquisicion'
grabador.writerows([(identificador, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for identificador, datos in datos_grabar.items()])
archivo.close()
conn.close()
#A